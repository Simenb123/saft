# app/parsers/saft_mva_termin_report.py
# -*- coding: utf-8 -*-
"""
MVA-termin-rapport (formelfri og "safe mode" for Excel):
  Faner:
    1) MVA-Term  – MVA-beløp per TaxCode per termin (T1..T6 pr år) + TOTAL-kolonne
    2) Grunnlag  – Grunnlag per TaxCode per termin (TaxBase* fra fil hvis mulig,
                   ellers MVA / (standard-sats/100)) + TOTAL-kolonne
    3) Avvik     – Alltid med. Viser koder der standard-sats != sats i filen
                   (TaxPercentageFile). Hvis ingen avvik -> en rad med "Ingen avvik".

Beløp            : MVA = DebitTaxAmount - CreditTaxAmount (fallback TaxAmount)
Grunnlag (primær): TaxBase/TaxBaseAmount/TaxableAmount hvis tilgjengelig
Grunnlag (sek.)  : MVA / (standard-sats/100) – rundes til 2 des

Dato             : PostingDate (fallback TransactionDate)
Filter           : TaxType in {"", "VAT", "MVA"} og TaxCountryRegion in {"", "NO", "NOR"}

Utdata:
  - csv/mva_termin_avstemming.csv
  - excel/mva_termin_avstemming.xlsx
"""
from __future__ import annotations
from pathlib import Path
from typing import Dict, Optional, Tuple, List
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
import csv
import re

# Vi foretrekker xlsxwriter (veldig robust). Hvis ikke tilgjengelig -> openpyxl.
try:
    import xlsxwriter  # type: ignore
    _HAS_XLSXWRITER = True
except Exception:
    _HAS_XLSXWRITER = False

try:
    import openpyxl  # type: ignore
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

SAFE_EXCEL_MODE = True  # skriv så "kjedelig" som mulig (ingen formler, ingen freeze/autofilter)

# ----------------- små hjelpere -----------------

def _as_dec_maybe(s: Optional[str]) -> Optional[Decimal]:
    if s is None:
        return None
    txt = s.strip()
    if not txt:
        return None
    txt = txt.replace(" ", "").rstrip("%").replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        try:
            return Decimal(str(float(txt)))
        except Exception:
            return None

def _to_dec(s: Optional[str]) -> Decimal:
    v = _as_dec_maybe(s)
    return v if v is not None else Decimal("0")

def _q2(v: Decimal) -> Decimal:
    return v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def _parse_date(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y%m%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    for fmt in ("%Y-%m", "%Y/%m"):
        try:
            d = datetime.strptime(s, fmt)
            return datetime(d.year, d.month, 1)
        except Exception:
            continue
    return None

def _term_from_month(month: int) -> int:
    return ((month - 1) // 2) + 1

def _read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))

def _pick_any(d: Dict[str, str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in d and (d[c] or "").strip():
            return d[c].strip()
    return None

def _sanitize_text(s: str) -> str:
    # Fjern kontrolltegn som Excel kan mislike i XML (ikke-trykkbare)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", s)


# ----------------- Offisielle standardnavn og -satser -----------------
# (Summene beregnes ALDRI fra satsene; satsene vises kun i metadata/avvik.)
_STD_NAMES: Dict[str, str] = {
    "0":  "Ingen merverdiavgiftsbehandling (anskaffelser)",
    "1":  "Kjøp av varer og tjenester med fradragsrett (høy sats)",
    "11": "Kjøp av varer og tjenester med fradragsrett (middels sats)",
    "12": "Kjøp av fisk og andre viltlevende marine ressurser (11,11 %)",
    "13": "Kjøp av varer og tjenester med fradragsrett (lav sats)",

    "3":  "Salg og uttak av varer og tjenester (høy sats)",
    "31": "Salg og uttak av varer og tjenester (middels sats)",
    "33": "Salg og uttak av varer og tjenester (lav sats)",
    "5":  "Salg og uttak av varer og tjenester fritatt for merverdiavgift (nullsats)",
    "6":  "Salg og uttak av varer og tjenester unntatt merverdiavgift",
    "7":  "Ingen merverdiavgiftsbehandling (inntekter)",
    "32": "Salg av fisk og andre viltlevende marine ressurser (11,11 %)",
    "51": "Salg av klimakvoter og gull (nullsats)",
    "52": "Salg av varer og tjenester til utlandet (nullsats)",

    "81": "Kjøp av varer fra utlandet med fradragsrett (høy sats)",
    "82": "Kjøp av varer fra utlandet uten fradragsrett (høy sats)",
    "83": "Kjøp av varer fra utlandet med fradragsrett (middels sats)",
    "84": "Kjøp av varer fra utlandet uten fradragsrett (middels sats)",
    "85": "Kjøp av varer fra utlandet, fritatt for merverdiavgift (nullsats)",
    "14": "Fradrag på mva betalt ved innførsel (høy sats)",
    "15": "Fradrag på mva betalt ved innførsel (middels sats)",
    "21": "Grunnlag ved innførsel av varer – høy sats",
    "22": "Grunnlag ved innførsel av varer – middels sats",

    "86": "Kjøp av tjenester fra utlandet med fradragsrett (høy sats)",
    "87": "Kjøp av tjenester fra utlandet uten fradragsrett (høy sats)",
    "88": "Kjøp av tjenester fra utlandet med fradragsrett (lav sats)",
    "89": "Kjøp av tjenester fra utlandet uten fradragsrett (lav sats)",

    "91": "Kjøp med omvendt avgiftsplikt i Norge (høy sats)",
    "92": "Kjøp med omvendt avgiftsplikt i Norge (middels sats)",
}

_STD_RATE: Dict[str, Decimal] = {
    "0": Decimal("0"),
    "1": Decimal("25"), "11": Decimal("15"), "12": Decimal("11.11"), "13": Decimal("12"),
    "3": Decimal("25"), "31": Decimal("15"), "33": Decimal("12"),
    "5": Decimal("0"), "6": Decimal("0"), "7": Decimal("0"), "32": Decimal("11.11"),
    "51": Decimal("0"), "52": Decimal("0"),
    "81": Decimal("25"), "82": Decimal("25"), "83": Decimal("15"), "84": Decimal("15"), "85": Decimal("0"),
    "14": Decimal("25"), "15": Decimal("15"), "21": Decimal("25"), "22": Decimal("15"),
    "86": Decimal("25"), "87": Decimal("25"), "88": Decimal("12"), "89": Decimal("12"),
    "91": Decimal("25"), "92": Decimal("25"),
}


# ----------------- metadata fra tax_table + mapping_tax -----------------

def _build_tax_meta(csv_dir: Path) -> Dict[str, Dict[str, str]]:
    """
    Meta pr TaxCode:
      - StandardTaxCode, StandardTaxCodeName (offisielt navn)
      - Description (fra filen)
      - TaxPercentage (standard)
      - TaxPercentageFile (fra filen) + TaxPercSource
    """
    meta: Dict[str, Dict[str, str]] = {}

    for r in _read_csv(csv_dir / "tax_table.csv"):
        code = (r.get("TaxCode") or r.get("Code") or "").strip()
        if not code:
            continue
        std  = (r.get("StandardTaxCode") or r.get("StandardCode") or "").strip()
        desc = (r.get("Description") or "").strip()
        pct  = _as_dec_maybe(_pick_any(r, ["TaxPercentage","TaxRate","Rate","Percent","Perc","Percentage"]))
        meta[code] = {
            "StandardTaxCode": std,
            "StandardTaxCodeName": _STD_NAMES.get(std, ""),
            "Description": desc,
            "TaxPercentage": f"{_STD_RATE.get(std, '')}",
            "TaxPercentageFile": f"{pct}" if pct is not None else "",
            "TaxPercSource": "tax_table" if pct is not None else "",
        }

    for r in _read_csv(csv_dir / "mapping_tax.csv"):
        code = (r.get("TaxCode") or r.get("Code") or r.get("TaxId") or "").strip()
        if not code:
            continue
        d = meta.setdefault(code, {})
        if not d.get("StandardTaxCode"):
            std = (r.get("StandardTaxCode") or r.get("StandardCode") or "").strip()
            d["StandardTaxCode"] = std
            d["StandardTaxCodeName"] = _STD_NAMES.get(std, "")
            d["TaxPercentage"] = f"{_STD_RATE.get(std, '')}"
        else:
            std = d.get("StandardTaxCode", "")
            d.setdefault("StandardTaxCodeName", _STD_NAMES.get(std, ""))
            d.setdefault("TaxPercentage", f"{_STD_RATE.get(std, '')}")
        if not d.get("Description"):
            d["Description"] = _pick_any(r, ["Description","TaxName","Name","Beskrivelse"]) or ""
        if not d.get("TaxPercentageFile"):
            pct = _as_dec_maybe(_pick_any(r, ["TaxPercentage","TaxRate","Rate","Percent","Perc","Percentage"]))
            if pct is not None:
                d["TaxPercentageFile"] = f"{pct}"
                d["TaxPercSource"] = "mapping_tax"
    return meta


# ----------------- Excel-skriving (formelfri) -----------------

def _write_xlsx_xlsxwriter(out_path: Path,
                           sheets: List[Tuple[str, List[str], List[List[str]]]]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(out_path))

    fmt_head = wb.add_format({"bold": True})
    fmt_num  = wb.add_format({"num_format": "#,##0.00;[Red]-#,##0.00"})
    fmt_pct  = wb.add_format({"num_format": "0.##"})
    fmt_txt  = wb.add_format({})

    for title, headers, table in sheets:
        ws = wb.add_worksheet(title)
        # Header
        for c, h in enumerate(headers):
            ws.write(0, c, _sanitize_text(h), fmt_head)
        # Data
        for r, row in enumerate(table, start=1):
            for c, v in enumerate(row):
                vv = "" if v is None else str(v)
                num = _as_dec_maybe(vv)
                if num is not None and headers[c] in ("TaxPercentage", "TaxPercentageFile", "DiffPct"):
                    ws.write_number(r, c, float(num), fmt_pct)
                elif num is not None and (headers[c].startswith("20") or headers[c] == "TOTAL" or headers[c] in ("TotalMVA","TotalGrunnlag")):
                    ws.write_number(r, c, float(num), fmt_num)
                else:
                    ws.write(r, c, _sanitize_text(vv), fmt_txt)
        # Kolonnebredder (enkelt og robust)
        for c in range(len(headers)):
            ws.set_column(c, c, 14)
        ws.set_column(0, 0, 12)   # TaxCode
        if len(headers) > 1: ws.set_column(1, 1, 14)   # StandardTaxCode
        if len(headers) > 2: ws.set_column(2, 2, 36)   # StandardTaxCodeName
        if len(headers) > 3: ws.set_column(3, 3, 36)   # Description

        # SAFE_EXCEL_MODE => ingen freeze/autofilter for å minimere "Repair"-risiko

    wb.close()


def _write_xlsx_openpyxl(out_path: Path,
                         sheets: List[Tuple[str, List[str], List[List[str]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Fjern default-ark
    wb.remove(wb.active)

    for title, headers, table in sheets:
        ws = wb.create_sheet(title=title[:31])  # Excel 31-char limit
        ws.append(headers)
        for row in table:
            ws.append([_sanitize_text("" if v is None else str(v)) for v in row])

        # Enkle bredder + tallformat
        widths = [12, 14, 36, 36] + [14] * max(0, len(headers) - 4)
        for i, w in enumerate(widths, start=1):
            if i <= len(headers):
                ws.column_dimensions[get_column_letter(i)].width = w

        # SAFE_EXCEL_MODE => ingen freeze/autofilter

    wb.save(out_path)


# ----------------- Hovedfunksjon -----------------

def make_mva_termin_report(csv_dir: Path, excel_dir: Optional[Path] = None) -> Optional[Path]:
    csv_dir = Path(csv_dir)
    tx_path = csv_dir / "transactions.csv"
    if not tx_path.exists():
        return None

    tax_meta = _build_tax_meta(csv_dir)

    # Aggreger pr (TaxCode, Year, Term)
    from collections import defaultdict
    mva = defaultdict(Decimal)   # (code, year, term) -> MVA-beløp
    base = defaultdict(Decimal)  # (code, year, term) -> grunnlag
    codes_seen: set[str] = set()
    term_keys: set[Tuple[int, int]] = set()

    base_candidates = [
        "TaxBase", "TaxBaseAmount", "TaxableAmount", "TaxBasis",
        "TaxableBase", "BaseAmount", "TaxAmountBase"
    ]

    with tx_path.open("r", encoding="utf-8", newline="") as f_in:
        reader = csv.DictReader(f_in)
        for row in reader:
            code = (row.get("TaxCode") or "").strip()
            if not code:
                continue

            # Filter: type & region
            tax_type = (row.get("TaxType") or "").strip().upper()
            if tax_type and tax_type not in ("VAT", "MVA"):
                continue
            region = (row.get("TaxCountryRegion") or "").strip().upper()
            if region and region not in ("", "NO", "NOR"):
                continue

            dt = _parse_date(row.get("PostingDate") or row.get("TransactionDate"))
            if not dt:
                continue
            year, term = dt.year, _term_from_month(dt.month)

            dta = _to_dec(row.get("DebitTaxAmount"))
            cta = _to_dec(row.get("CreditTaxAmount"))
            tax_amt = dta - cta
            if tax_amt == 0:
                tax_amt = _to_dec(row.get("TaxAmount"))
            if tax_amt == 0:
                continue

            # Grunnlag fra fil?
            base_val = Decimal("0")
            for cand in base_candidates:
                v = _as_dec_maybe(row.get(cand))
                if v is not None:
                    base_val = v
                    break
            # Hvis ikke -> beregn fra standard-sats
            if base_val == 0:
                std = tax_meta.get(code, {}).get("StandardTaxCode", "")
                rate = _STD_RATE.get(std)
                if rate and rate != 0:
                    base_val = _q2(tax_amt / (rate / Decimal("100")))

            mva[(code, year, term)]  += tax_amt
            base[(code, year, term)] += base_val
            codes_seen.add(code)
            term_keys.add((year, term))

    if not codes_seen:
        return None

    # Kolonner (år/termin)
    term_list = sorted(term_keys)
    # Bruk ASCII bindestrek for maksimal Excel-kompatibilitet
    month_spans = {1: "Jan-Feb", 2: "Mar-Apr", 3: "Mai-Jun", 4: "Jul-Aug", 5: "Sep-Okt", 6: "Nov-Des"}

    base_headers = [
        "TaxCode", "StandardTaxCode", "StandardTaxCodeName",
        "Description", "TaxPercentage", "TaxPercentageFile", "TaxPercSource"
    ]
    term_headers = [f"{yy}-T{tt} ({month_spans.get(tt,'')})" for (yy, tt) in term_list]
    headers = base_headers + term_headers + ["TOTAL"]
    term_start_idx = len(base_headers)

    # Sorter rader: StandardTaxCode, deretter TaxCode
    def _row_sort_key(c: str):
        return (tax_meta.get(c, {}).get("StandardTaxCode", ""), c)
    row_codes = sorted(codes_seen, key=_row_sort_key)

    # Tabeller + sum-rader
    table_mva: List[List[str]] = []
    table_base: List[List[str]] = []
    sum_row_mva = ["SUM", "", "", "", "", "", ""]
    sum_row_base = ["SUM", "", "", "", "", "", ""]

    for code in row_codes:
        meta = tax_meta.get(code, {})
        row_meta = [
            code,
            meta.get("StandardTaxCode", ""),
            meta.get("StandardTaxCodeName", ""),
            meta.get("Description", ""),
            meta.get("TaxPercentage", ""),
            meta.get("TaxPercentageFile", ""),
            meta.get("TaxPercSource", ""),
        ]
        # MVA
        vals_mva: List[Decimal] = []
        for (yy, tt) in term_list:
            vals_mva.append(mva.get((code, yy, tt), Decimal("0")))
        tot_mva = sum(vals_mva, Decimal("0"))
        table_mva.append(row_meta + [f"{_q2(v)}" for v in vals_mva] + [f"{_q2(tot_mva)}"])

        # Grunnlag
        vals_base: List[Decimal] = []
        for (yy, tt) in term_list:
            vals_base.append(base.get((code, yy, tt), Decimal("0")))
        tot_base = sum(vals_base, Decimal("0"))
        table_base.append(row_meta + [f"{_q2(v)}" for v in vals_base] + [f"{_q2(tot_base)}"])

    # SUM-kolonner
    for (yy, tt) in term_list:
        s_mva  = sum((mva.get((c, yy, tt), Decimal("0")) for c in row_codes),  Decimal("0"))
        s_base = sum((base.get((c, yy, tt), Decimal("0")) for c in row_codes), Decimal("0"))
        sum_row_mva.append(f"{_q2(s_mva)}")
        sum_row_base.append(f"{_q2(s_base)}")
    sum_row_mva.append(f"{_q2(sum((_as_dec_maybe(x) or Decimal('0')) for x in sum_row_mva[term_start_idx:]))}")
    sum_row_base.append(f"{_q2(sum((_as_dec_maybe(x) or Decimal('0')) for x in sum_row_base[term_start_idx:]))}")

    # Avvik-tabell (alltid)
    headers_dev = [
        "TaxCode", "StandardTaxCode", "StandardTaxCodeName",
        "TaxPercentage", "TaxPercentageFile", "DiffPct",
        "TotalMVA", "TotalGrunnlag"
    ]
    table_dev: List[List[str]] = []
    for code in row_codes:
        meta = tax_meta.get(code, {})
        std_pct  = _as_dec_maybe(meta.get("TaxPercentage")) or Decimal("0")
        file_pct = _as_dec_maybe(meta.get("TaxPercentageFile") or "")
        if file_pct is None:
            continue
        diff = std_pct - file_pct
        if abs(diff) >= Decimal("0.01"):
            tot_mva  = sum((mva.get((code, yy, tt),  Decimal("0")) for (yy, tt) in term_list), Decimal("0"))
            tot_base = sum((base.get((code, yy, tt), Decimal("0")) for (yy, tt) in term_list), Decimal("0"))
            table_dev.append([
                code, meta.get("StandardTaxCode",""), meta.get("StandardTaxCodeName",""),
                f"{_q2(std_pct)}", f"{_q2(file_pct)}", f"{_q2(diff)}",
                f"{_q2(tot_mva)}", f"{_q2(tot_base)}"
            ])
    if not table_dev:
        table_dev.append(["Ingen avvik", "", "", "", "", "", "", ""])

    # CSV
    out_csv = csv_dir / "mva_termin_avstemming.csv"
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(headers)
        for row in table_mva:
            w.writerow(row)
        w.writerow(sum_row_mva)

    # Excel
    if excel_dir is None:
        excel_dir = csv_dir.parent / "excel"
    out_xlsx = Path(excel_dir) / "mva_termin_avstemming.xlsx"

    try:
        sheets = [
            ("MVA-Term", headers, table_mva + [sum_row_mva]),
            ("Grunnlag", headers, table_base + [sum_row_base]),
            ("Avvik", headers_dev, table_dev),
        ]
        if _HAS_XLSXWRITER:
            _write_xlsx_xlsxwriter(out_xlsx, sheets)
        elif _HAS_OPENPYXL:
            _write_xlsx_openpyxl(out_xlsx, sheets)
        else:
            return None
    except Exception:
        return None

    return out_xlsx
