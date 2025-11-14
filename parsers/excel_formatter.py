# -*- coding: utf-8 -*-
"""
excel_formatter.py
------------------
Setter visningsformat i Excel-rapporter:
 - Tall: tusenskiller og 2 desimaler  ->  "#,##0.00"
 - Dato: norsk  dd.mm.yyyy
 - Fryser første rad (header)

Formatet er KUN visning i Excel. CSV endres ikke.

Brukes av runner/GUI slik:
    from app.parsers import excel_formatter
    excel_formatter.format_all(excel_dir, verbose=False)

Heuristikkene er utvidet til å treffe Trial Balance-kolonner som IB/Movement/UB.
"""
from __future__ import annotations

from pathlib import Path
from typing import List, Dict
import re
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


# --- Heuristikk for kolonnenavn (engelsk + norsk) ---------------------------

_AMOUNT_KEYS = (
    # generelt
    "amount", "debit", "credit", "balance", "saldo", "beløp",
    "avgift", "mva", "netamount", "grossamount", "baseamount", "openamount",
    # trial balance spesifikt
    "ib", "openingbalance", "opening_balance", "opening",
    "movement", "netchange", "change", "periodmovement", "period_debit", "period_credit",
    "ub", "closingbalance", "closing_balance", "closing", "endbalance", "endingbalance",
)

_DATE_KEYS = (
    "date", "postingdate", "transactiondate", "invoicedate", "duedate",
    "documentdate", "journaldate", "entrydate",
    "dato", "bilagsdato", "posteringsdato", "forfallsdato", "fakturadato",
)

# ISO-datoer vi trygt kan konvertere til ekte dato
_ISO_DATE_RE = re.compile(r"^\s*(\d{4})-(\d{2})-(\d{2})(?:[T\s].*)?\s*$")


def _classify_header(header: str) -> str:
    if not header:
        return "other"
    h = header.strip().lower()
    for k in _DATE_KEYS:
        if k in h:
            return "date"
    for k in _AMOUNT_KEYS:
        if k in h:
            return "amount"
    return "other"


def _iter_header(ws, header_row: int = 1) -> List[str]:
    headers: List[str] = []
    for cell in ws[header_row]:
        v = cell.value
        headers.append("" if v is None else str(v))
    return headers


def _format_ws(ws, header_row: int = 1) -> Dict[str, int]:
    """Formater ett ark og frys første rad."""
    max_row = ws.max_row
    ws.freeze_panes = f"A{header_row+1}"  # alltid

    if max_row < header_row + 1:
        return {"amount_cols": 0, "date_cols": 0}

    headers = _iter_header(ws, header_row)
    amount_cols, date_cols = [], []

    for idx, name in enumerate(headers, start=1):
        kind = _classify_header(name)
        if kind == "amount":
            amount_cols.append(idx)
        elif kind == "date":
            date_cols.append(idx)

    # Tallformat: tusenskiller og 2 desimaler
    for c in amount_cols:
        for cell in ws.iter_rows(min_row=header_row+1, max_row=max_row,
                                 min_col=c, max_col=c):
            cell[0].number_format = "#,##0.00"

    # Datoformat: dd.mm.yyyy (konverter ISO-strenger forsiktig)
    for c in date_cols:
        for cell in ws.iter_rows(min_row=header_row+1, max_row=max_row,
                                 min_col=c, max_col=c):
            cell_obj = cell[0]
            v = cell_obj.value
            if isinstance(v, (datetime, date)):
                cell_obj.number_format = "dd.mm.yyyy"
                continue
            if isinstance(v, str):
                m = _ISO_DATE_RE.match(v)
                if m:
                    y, mth, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    try:
                        cell_obj.value = date(y, mth, d)
                    except Exception:
                        pass
            cell_obj.number_format = "dd.mm.yyyy"

    return {"amount_cols": len(amount_cols), "date_cols": len(date_cols)}


def format_workbook(xlsx_path: Path, *, header_row: int = 1, verbose: bool = False) -> int:
    if load_workbook is None:
        if verbose:
            print(f"[excel_formatter] openpyxl er ikke installert – hopper over {xlsx_path}")
        return 0

    wb = load_workbook(str(xlsx_path))
    total = 0
    for ws in wb.worksheets:
        stats = _format_ws(ws, header_row=header_row)
        total += 1
        if verbose:
            print(f"[excel_formatter] {xlsx_path.name} :: {ws.title}  "
                  f"(amount_cols={stats['amount_cols']}, date_cols={stats['date_cols']})")
    wb.save(str(xlsx_path))
    return total


def format_all(excel_dir: Path, *, header_row: int = 1, verbose: bool = False) -> int:
    excel_dir = Path(excel_dir)
    if not excel_dir.exists():
        return 0

    count = 0
    for p in sorted(excel_dir.glob("*.xlsx")):
        try:
            n = format_workbook(p, header_row=header_row, verbose=verbose)
            if n > 0:
                count += 1
        except Exception as e:
            if verbose:
                print(f"[excel_formatter] Skipping {p.name}: {e!r}")
    return count
