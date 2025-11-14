# -*- coding: utf-8 -*-
"""
GL monthly – Pivot_Period (IB | 1..12 | UB)

v11 (robust kontonavn + totals + no SettingWithCopy):
- Kun IB-konti med verdi (IB != 0) suppleres hvis de ikke finnes i GL.
- Kontonavn fylles for IB-kun-konti fra kombinasjon av GL-navn (hvis finnes) og master-navn
  (accounts.csv eller trial_balance.xlsx). GL-navn prioriteres, master brukes som fallback per konto.
- Join-keys er alltid ['AccountID'] (+ ['Year'] ved flerårig fil).
- Serietotaler 1xxx..9xxx + SUM-rad appendes nederst (både trial_balance.xlsx og gl_monthly.xlsx).
- Norsk regnskapsformat på tall.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, List, Tuple
import pandas as pd, numpy as np, re

ACCOUNTING_FORMAT = '_-* # ##0,00_-;_-* (# ##0,00)_-;_-* "-"_-;_-@_-'
DATE_FORMAT = 'yyyy-mm-dd'
SHEET_NAME = "Pivot_Period"

# ---------- I/O helpers ----------

def _read_csv_safe(path: Path | str, dtype: str | dict = "str") -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(p, dtype=dtype, encoding="utf-8-sig", sep=None, engine="python")
    except Exception:
        try:
            return pd.read_csv(p, dtype=dtype, encoding="utf-8-sig", sep=";")
        except Exception:
            return pd.read_csv(p, dtype=dtype, encoding="utf-8")

def _find_csv(base: Path, name: str) -> Optional[Path]:
    for cand in [Path(base) / name, Path(base).parent / "csv" / name, Path(base) / "csv" / name]:
        if Path(cand).exists():
            return Path(cand)
    return None

def _xlsx_writer(path: Path) -> pd.ExcelWriter:
    return pd.ExcelWriter(
        str(path),
        engine="xlsxwriter",
        datetime_format=DATE_FORMAT,
        engine_kwargs={"options": {"strings_to_urls": False, "strings_to_numbers": False, "strings_to_formulas": False}}
    )

def _apply_formats_xlsxwriter(xw, sheet: str, df: pd.DataFrame, numeric_cols: Optional[List[str]] = None):
    try:
        ws = xw.sheets[sheet]
        book = xw.book
        fmt_num = book.add_format({"num_format": ACCOUNTING_FORMAT})
        fmt_dt  = book.add_format({"num_format": DATE_FORMAT})
        cols = list(df.columns)
        head = [len(str(c)) for c in cols]
        sample = df.head(500)
        for i, c in enumerate(cols):
            try:
                m = int(min(sample[c].astype(str).map(len).max(), 60))
            except Exception:
                m = 8
            w = max(10, min(60, max(head[i], m) + 2))
            if (numeric_cols and c in numeric_cols) or (pd.api.types.is_numeric_dtype(df[c])):
                ws.set_column(i, i, w, fmt_num)
            elif "Date" in str(c):
                ws.set_column(i, i, w, fmt_dt)
            else:
                ws.set_column(i, i, w)
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, 0, len(cols) - 1)
    except Exception:
        pass

def _apply_formats_openpyxl(path: Path, sheet: str, df: pd.DataFrame, numeric_cols: Optional[List[str]] = None):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        num_cols = numeric_cols or [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        name_to_idx = {cell.value: idx+1 for idx, cell in enumerate(ws[1]) if cell.value}
        for name in num_cols:
            if name not in name_to_idx:
                continue
            col_idx = name_to_idx[name]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                row[0].number_format = '# ##0,00_);(# ##0,00);"-"'
        wb.save(path)
    except Exception:
        pass

# ---------- Common helpers ----------

def _norm_id(s: pd.Series) -> pd.Series:
    """Normaliser kontonr til streng, trim og fjern '.0' fra Excel-flytall."""
    out = s.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    return out

def _first_digit(s: str) -> Optional[str]:
    m = re.search(r"\d", str(s) if s is not None else "")
    return m.group(0) if m else None

def _parse_date(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in ["PostingDate","TransactionDate","Date"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce")
    if "Date" not in out.columns:
        out["Date"] = out.get("PostingDate").fillna(out.get("TransactionDate"))
    return out

def _to_num(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
    return out

# ---------- Accounts (navn + IB) ----------

_ACCOUNT_ID_SYNONYMS = ["AccountID","Account","AccountNumber","Number","Konto","Kontonummer","Kontonr"]
_ACCOUNT_NAME_SYNONYMS = ["AccountDescription","AccountName","Description","Name","Kontonavn","Navn","Beskrivelse"]

def _pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = {c.lower(): c for c in df.columns}
    for n in candidates:
        if n.lower() in cols:
            return cols[n.lower()]
    # contains search
    for c in df.columns:
        cl = c.lower()
        for n in candidates:
            if n.lower() in cl:
                return c
    return None

def _read_account_master(out_dir: Path) -> pd.DataFrame:
    """Returnerer AccountID + AccountDescription fra accounts.csv eller trial_balance.xlsx"""
    # accounts.csv
    p = _find_csv(out_dir, "accounts.csv")
    if p:
        df = _read_csv_safe(p, dtype=str)
        aid = _pick_col(df, _ACCOUNT_ID_SYNONYMS)
        anm = _pick_col(df, _ACCOUNT_NAME_SYNONYMS)
        if aid:
            res = pd.DataFrame()
            res["AccountID"] = _norm_id(df[aid])
            res["AccountDescription"] = df[anm].astype(str).fillna("") if anm else ""
            return res.drop_duplicates(subset=["AccountID"])

    # trial_balance.xlsx
    xls = (out_dir.parent / "excel" / "trial_balance.xlsx")
    if xls.exists():
        try:
            sheets = pd.read_excel(xls, sheet_name=None, dtype=str)
            for _, d in sheets.items():
                aid = _pick_col(d, _ACCOUNT_ID_SYNONYMS)
                anm = _pick_col(d, _ACCOUNT_NAME_SYNONYMS)
                if aid and anm:
                    res = pd.DataFrame()
                    res["AccountID"] = _norm_id(d[aid])
                    res["AccountDescription"] = d[anm].astype(str).fillna("")
                    return res.drop_duplicates(subset=["AccountID"])
        except Exception:
            pass

    return pd.DataFrame(columns=["AccountID","AccountDescription"])

def _derive_ib_from_accounts(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["AccountID","IB"])
    acc_col = _pick_col(df, _ACCOUNT_ID_SYNONYMS)
    if not acc_col:
        return pd.DataFrame(columns=["AccountID","IB"])
    d_col = _pick_col(df, ["OpeningDebit","OBDebit","OpeningBalanceDebit","DebitOpening","Opening_Debit","Åpningsdeb"])
    c_col = _pick_col(df, ["OpeningCredit","OBCredit","OpeningBalanceCredit","CreditOpening","Opening_Credit","Åpningskre"])
    ib_col = _pick_col(df, ["IB","OpeningBalance","InitialBalance","BeginBalance","BalanceBroughtForward"])

    out = pd.DataFrame()
    out["AccountID"] = _norm_id(df[acc_col])

    if d_col and c_col:
        d = pd.to_numeric(df[d_col], errors="coerce").fillna(0.0)
        c = pd.to_numeric(df[c_col], errors="coerce").fillna(0.0)
        out["IB"] = d - c
    elif ib_col:
        out["IB"] = pd.to_numeric(df[ib_col], errors="coerce").fillna(0.0)
    else:
        out["IB"] = 0.0

    return out.groupby("AccountID", as_index=False)["IB"].sum()

def _read_ib_and_master(out_dir: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Returnerer (ib_df, master_df). master_df har AccountDescription om tilgjengelig."""
    # primary: accounts.csv
    p = _find_csv(out_dir, "accounts.csv")
    if p:
        df = _read_csv_safe(p, dtype=str)
        ib = _derive_ib_from_accounts(df)
        master = _read_account_master(out_dir)  # leser fra samme fil evt.
        return ib, master

    # fallback: trial_balance.xlsx
    xls = (out_dir.parent / "excel" / "trial_balance.xlsx")
    if xls.exists():
        try:
            sheets = pd.read_excel(xls, sheet_name=None, dtype=str)
            # finn ark med AccountID + (IB eller Opening…)
            for _, d in sheets.items():
                has_id = _pick_col(d, _ACCOUNT_ID_SYNONYMS) is not None
                has_ib = _pick_col(d, ["IB","OpeningBalance"]) is not None or \
                         _pick_col(d, ["OpeningDebit"]) is not None or \
                         _pick_col(d, ["OpeningCredit"]) is not None
                if has_id and has_ib:
                    ib = _derive_ib_from_accounts(d)
                    master = _read_account_master(out_dir)
                    return ib, master
        except Exception:
            pass
    # last resort
    return pd.DataFrame(columns=["AccountID","IB"]), _read_account_master(out_dir)

# ---------- Pivot builder ----------

def _pivot_period_with_ib(out_dir: Path, df: pd.DataFrame) -> pd.DataFrame:
    df = _parse_date(df)
    df = _to_num(df, ["Debit","Credit","Amount","Year","Period"])
    if "Amount" not in df.columns:
        df["Amount"] = df.get("Debit", 0.0) - df.get("Credit", 0.0)

    if "Period" in df.columns and pd.to_numeric(df["Period"], errors="coerce").notna().any():
        df["PeriodNum"] = pd.to_numeric(df["Period"], errors="coerce").astype("Int64")
    else:
        df["PeriodNum"] = pd.to_datetime(df["Date"], errors="coerce").dt.month.astype("Int64")

    df = df[(df["PeriodNum"] >= 1) & (df["PeriodNum"] <= 12)].copy()

    has_year = "Year" in df.columns and df["Year"].astype(str).nunique(dropna=True) > 1
    year_col = "Year" if has_year else None

    # JOIN KEYS: alltid AccountID (+ Year hvis flere år)
    if "AccountID" in df.columns:
        df.loc[:, "AccountID"] = _norm_id(df["AccountID"])
    join_keys = ["AccountID"] + ([year_col] if year_col else [])

    # Bevegelse pr konto/periode
    if not df.empty:
        g = df.groupby(join_keys + ["PeriodNum"], dropna=False)["Amount"].sum().reset_index()
        piv_mov = g.pivot_table(index=join_keys, columns="PeriodNum", values="Amount", aggfunc="sum", fill_value=0.0).reset_index()
    else:
        piv_mov = pd.DataFrame(columns=join_keys + list(range(1,13)))

    # IB + kontonavn
    ib, master = _read_ib_and_master(out_dir)  # AccountID/IB og AccountID/AccountDescription
    if not ib.empty and "AccountID" in ib.columns:
        ib.loc[:, "AccountID"] = _norm_id(ib["AccountID"])
    if not master.empty and "AccountID" in master.columns:
        master.loc[:, "AccountID"] = _norm_id(master["AccountID"])

    # Indeks (GL-konti + IB-konti med verdi)
    idx_df = df.drop_duplicates(subset=join_keys)[join_keys].copy()
    if not ib.empty:
        ib.loc[:, "IB"] = pd.to_numeric(ib["IB"], errors="coerce").fillna(0.0)
        non_zero_ib = ib[np.isfinite(ib["IB"]) & (ib["IB"] != 0)]
        missing = non_zero_ib[~non_zero_ib["AccountID"].isin(idx_df["AccountID"].astype(str))][["AccountID"]].copy()
        if year_col:
            for yr in sorted(df[year_col].dropna().unique().tolist() or [None]):
                blk = missing.copy()
                if blk.empty:
                    continue
                blk.loc[:, year_col] = yr
                idx_df = pd.concat([idx_df, blk], ignore_index=True).drop_duplicates()
        else:
            idx_df = pd.concat([idx_df, missing], ignore_index=True).drop_duplicates()

    # Merge bevegelse + IB
    piv_mov.columns = [str(c) if isinstance(c, int) else c for c in piv_mov.columns]
    res = idx_df.merge(piv_mov, on=join_keys, how="left")
    for c in range(1,13):
        col = str(c)
        if col not in res.columns:
            res.loc[:, col] = 0.0
    for c in [str(c) for c in range(1,13)]:
        res.loc[:, c] = pd.to_numeric(res[c], errors="coerce").fillna(0.0)

    if not ib.empty:
        res = res.merge(ib, on="AccountID", how="left")
    if "IB" not in res.columns:
        res.loc[:, "IB"] = 0.0
    res.loc[:, "IB"] = pd.to_numeric(res["IB"], errors="coerce").fillna(0.0)

    # Kontonavn – GL-navn først, master som fallback
    desc_gl = pd.DataFrame()
    if "AccountDescription" in df.columns:
        tmp = df[["AccountID","AccountDescription"]].dropna()
        tmp = tmp[tmp["AccountID"].astype(str).str.strip() != ""]
        if not tmp.empty:
            tmp = tmp.copy()
            tmp.loc[:, "AccountID"] = _norm_id(tmp["AccountID"])
            desc_gl = tmp.drop_duplicates(subset=["AccountID"])
    desc_m = pd.DataFrame()
    if not master.empty:
        keep = ["AccountID","AccountDescription"]
        present = [c for c in keep if c in master.columns]
        if present == keep:
            desc_m = master[keep].dropna()
            desc_m = desc_m[desc_m["AccountID"].astype(str).str.strip() != ""]
            if not desc_m.empty:
                desc_m = desc_m.drop_duplicates(subset=["AccountID"])

    desc_gl_map = desc_gl.set_index("AccountID")["AccountDescription"] if not desc_gl.empty else pd.Series(dtype=object)
    desc_m_map  = desc_m.set_index("AccountID")["AccountDescription"]  if not desc_m.empty else pd.Series(dtype=object)

    res.loc[:, "AccountDescription"] = res["AccountID"].map(desc_gl_map)
    missing_mask = res["AccountDescription"].isna() | (res["AccountDescription"].astype(str).str.strip() == "")
    if not desc_m_map.empty:
        res.loc[missing_mask, "AccountDescription"] = res.loc[missing_mask, "AccountID"].map(desc_m_map)
    res["AccountDescription"] = res["AccountDescription"].fillna("")

    res.loc[:, "UB"] = res["IB"] + res[[str(i) for i in range(1,13)]].sum(axis=1)

    ordered = ["AccountID","AccountDescription"] + ([year_col] if year_col else []) + ["IB"] + [str(i) for i in range(1,13)] + ["UB"]
    ordered = [c for c in ordered if c in res.columns]
    res = res[ordered]

    sort_cols = [c for c in ["AccountID", year_col] if c in res.columns]
    if sort_cols:
        res = res.sort_values(sort_cols, kind="mergesort")
    return res

# ---------- Totals & series ----------

def _numeric_cols(df: pd.DataFrame) -> List[str]:
    return ["IB"] + [str(i) for i in range(1,13)] + ["UB"]

def _append_totals_series_openpyxl(path: Path, sheet: str, df: pd.DataFrame):
    """Legg SUM-rad + serietotaler (1xxx..9xxx) i et eksisterende ark (openpyxl)."""
    try:
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        name_to_idx = {cell.value: idx+1 for idx, cell in enumerate(ws[1]) if cell.value}
        r = ws.max_row + 2
        bold = Font(bold=True)
        num_cols = [c for c in _numeric_cols(df) if c in name_to_idx]

        ws.cell(r, 1, "SUM").font = bold
        for c in num_cols:
            v = float(pd.to_numeric(df[c], errors="coerce").fillna(0.0).sum())
            cell = ws.cell(r, name_to_idx[c], v)
            cell.number_format = '# ##0,00_);(# ##0,00);"-"'
            cell.font = bold
        r += 2

        if "AccountID" in df.columns:
            fd = df["AccountID"].astype(str).map(_first_digit)
            for d in list("123456789"):
                mask = fd == d
                if not mask.any():
                    continue
                ws.cell(r, 1, f"SUM {d}xxx").font = bold
                for c in num_cols:
                    v = float(pd.to_numeric(df.loc[mask, c], errors="coerce").fillna(0.0).sum())
                    cell = ws.cell(r, name_to_idx[c], v)
                    cell.number_format = '# ##0,00_);(# ##0,00);"-"'
                    cell.font = bold
                r += 1
        wb.save(path)
    except Exception:
        pass

def _write_totals_series_xlsxwriter(xw, sheet: str, df: pd.DataFrame):
    try:
        ws = xw.sheets[sheet]
        book = xw.book
        fmt = book.add_format({"num_format": ACCOUNTING_FORMAT, "bold": True})
        fmt_lbl = book.add_format({"bold": True})
        rows = len(df)
        r = 1 + rows + 1
        ws.write(r, 0, "SUM", fmt_lbl)
        for i, col in enumerate(df.columns):
            if pd.api.types.is_numeric_dtype(df[col]) or col in _numeric_cols(df):
                cidx = list(df.columns).index(col)
                val = float(pd.to_numeric(df[col], errors="coerce").fillna(0.0).sum())
                ws.write_number(r, cidx, val, fmt)
        r += 2
        if "AccountID" in df.columns:
            fd = df["AccountID"].astype(str).map(_first_digit)
            for d in list("123456789"):
                mask = fd == d
                if not mask.any():
                    continue
                ws.write(r, 0, f"SUM {d}xxx", fmt_lbl)
                for i, col in enumerate(df.columns):
                    if (pd.api.types.is_numeric_dtype(df[col]) or col in _numeric_cols(df)):
                        cidx = list(df.columns).index(col)
                        val = float(pd.to_numeric(df.loc[mask, col], errors="coerce").fillna(0.0).sum())
                        ws.write_number(r, cidx, val, fmt)
                r += 1
    except Exception:
        pass

# ---------- Public API ----------

def make_gl_monthly(out_dir: Path) -> Path:
    """Bygg Pivot_Period og legg arket i trial_balance.xlsx + gl_monthly.xlsx (med SUM + serietotaler)."""
    out_dir = Path(out_dir)
    p = _find_csv(out_dir, "transactions.csv")
    if not p:
        raise FileNotFoundError("transactions.csv mangler")
    df = _read_csv_safe(p, dtype=str)
    if df.empty:
        raise ValueError("transactions.csv er tom")

    piv_per = _pivot_period_with_ib(out_dir, df)

    excel_dir = out_dir.parent / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)

    # 1) trial_balance.xlsx
    tb_path = excel_dir / "trial_balance.xlsx"
    try:
        if tb_path.exists():
            with pd.ExcelWriter(tb_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                piv_per.to_excel(xw, index=False, sheet_name=SHEET_NAME)
        else:
            with _xlsx_writer(tb_path) as xw:
                piv_per.to_excel(xw, index=False, sheet_name=SHEET_NAME)
                _apply_formats_xlsxwriter(xw, SHEET_NAME, piv_per, numeric_cols=_numeric_cols(piv_per))
        _apply_formats_openpyxl(tb_path, SHEET_NAME, piv_per, numeric_cols=_numeric_cols(piv_per))
        _append_totals_series_openpyxl(tb_path, SHEET_NAME, piv_per)
    except Exception:
        with _xlsx_writer(tb_path) as xw:
            piv_per.to_excel(xw, index=False, sheet_name=SHEET_NAME)
            _apply_formats_xlsxwriter(xw, SHEET_NAME, piv_per, numeric_cols=_numeric_cols(piv_per))
            _write_totals_series_xlsxwriter(xw, SHEET_NAME, piv_per)

    # 2) gl_monthly.xlsx (inntil vi bekrefter TrialBalance)
    gl_path = excel_dir / "gl_monthly.xlsx"
    with _xlsx_writer(gl_path) as xw:
        piv_per.to_excel(xw, index=False, sheet_name=SHEET_NAME)
        _apply_formats_xlsxwriter(xw, SHEET_NAME, piv_per, numeric_cols=_numeric_cols(piv_per))
        _write_totals_series_xlsxwriter(xw, SHEET_NAME, piv_per)

    print(f"[excel] Oppdatert trial_balance.xlsx ({SHEET_NAME}) + gl_monthly.xlsx med SUM og 1xxx..9xxx serietotaler")
    return tb_path

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Legg GL Pivot_Period i trial_balance.xlsx + gl_monthly.xlsx med SUM og serietotaler")
    p.add_argument("--out", dest="out_dir", required=True)
    args = p.parse_args()
    make_gl_monthly(Path(args.out_dir))
