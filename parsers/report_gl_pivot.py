# app/parsers/report_gl_pivot.py
# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path
from typing import Optional, List
import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception:
    load_workbook = None
    def get_column_letter(i: int) -> str:  # type: ignore
        return "A"

def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, dtype=str, encoding="utf-8-sig", sep=None, engine="python")
    except Exception:
        try:
            return pd.read_csv(path, dtype=str, encoding="utf-8-sig", sep=";")
        except Exception:
            return pd.read_csv(path, dtype=str, encoding="utf-8")

def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0.0)

def _auto_cols(ws, widths):
    try:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    except Exception:
        pass

def _numfmt(cell) -> None:
    try:
        cell.number_format = "# ##0,00;[Red]-# ##0,00"
    except Exception:
        pass

def _month_range(dfrom: Optional[pd.Timestamp], dto: Optional[pd.Timestamp], months_found: List[str]) -> List[str]:
    if dfrom is None or dto is None:
        return sorted(set(months_found))
    yms = []
    cur = pd.Period(dfrom, freq="M")
    end = pd.Period(dto, freq="M")
    while cur <= end:
        yms.append(str(cur))
        cur = cur + 1
    return yms

def attach_gl_pivot_sheet(outdir: Path,
                          date_from: Optional[str] = None,
                          date_to: Optional[str] = None,
                          sheet_name: str = "GL_Pivot") -> Path:
    """
    Legger til fane 'GL_Pivot' i trial_balance.xlsx (best effort).
    Kolonner: AccountID | AccountDescription | IB | YYYY-MM ... | UB
    """
    outdir = Path(outdir)
    tb_path = outdir / "trial_balance.xlsx"
    if not tb_path.exists() or load_workbook is None:
        return tb_path

    tx_path = outdir / "transactions.csv"
    acc_path = outdir / "accounts.csv"
    tx = _read_csv(tx_path)
    if tx.empty:
        return tb_path

    for c in ["PostingDate", "TransactionDate"]:
        if c in tx.columns:
            tx[c] = pd.to_datetime(tx[c], errors="coerce")
    if "PostingDate" in tx.columns:
        tx["Date"] = tx["PostingDate"].fillna(tx.get("TransactionDate"))
    else:
        tx["Date"] = tx.get("TransactionDate")
    if "Date" not in tx.columns or tx["Date"].isna().all():
        return tb_path

    for c in ["Debit", "Credit"]:
        if c in tx.columns:
            tx[c] = _to_num(tx[c])
    tx["Amount"] = tx.get("Debit", 0.0) - tx.get("Credit", 0.0)
    if "AccountID" in tx.columns:
        tx["AccountID"] = tx["AccountID"].fillna("").astype(str).str.strip()

    dfrom = pd.to_datetime(date_from) if date_from else None
    dto = pd.to_datetime(date_to) if date_to else None

    acc = _read_csv(acc_path)
    ib = pd.Series(dtype="float64")
    name_map = pd.Series(dtype="object")
    if not acc.empty and "AccountID" in acc.columns:
        acc["AccountID"] = acc["AccountID"].fillna("").astype(str).str.strip()
        if "AccountDescription" in acc.columns:
            name_map = acc.set_index("AccountID")["AccountDescription"]
        if "OpeningDebit" in acc.columns and "OpeningCredit" in acc.columns:
            ib = _to_num(acc["OpeningDebit"]) - _to_num(acc["OpeningCredit"])
            ib.index = acc.index
            ib = pd.Series(ib.values, index=acc["AccountID"]).groupby(level=0).sum()
        elif "OpeningBalance" in acc.columns:
            ib = pd.Series(_to_num(acc["OpeningBalance"]).values, index=acc["AccountID"]).groupby(level=0).sum()

    if ib.empty:
        pre = tx if dfrom is None else tx.loc[tx["Date"] < dfrom]
        if not pre.empty:
            ib = pre.groupby("AccountID")["Amount"].sum()

    if dfrom is not None:
        tx = tx.loc[tx["Date"] >= dfrom]
    if dto is not None:
        tx = tx.loc[tx["Date"] <= dto]
    if tx.empty and ib.empty:
        return tb_path

    tx["YYYYMM"] = tx["Date"].dt.to_period("M").astype(str)
    piv = tx.groupby(["AccountID", "YYYYMM"])["Amount"].sum().unstack(fill_value=0.0)

    months_found = list(piv.columns.astype(str)) if not piv.empty else []
    months = _month_range(dfrom, dto, months_found)
    for m in months:
        if m not in piv.columns:
            piv[m] = 0.0
    piv = piv[months].reset_index()

    if not name_map.empty:
        piv["AccountDescription"] = piv["AccountID"].map(name_map).fillna("")
        cols = ["AccountID", "AccountDescription"] + months
        piv = piv.reindex(columns=cols)
    else:
        piv["AccountDescription"] = ""
        piv = piv[["AccountID", "AccountDescription"] + months]

    if not ib.empty:
        ib_series = piv["AccountID"].map(ib).fillna(0.0)
    else:
        ib_series = pd.Series(0.0, index=piv.index)
    piv.insert(2, "IB", ib_series.values)

    month_cols = months
    for c in month_cols:
        piv[c] = _to_num(piv[c])
    piv["UB"] = piv["IB"] + piv[month_cols].sum(axis=1)

    out = piv[["AccountID", "AccountDescription", "IB"] + month_cols + ["UB"]].sort_values("AccountID")

    wb = load_workbook(tb_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = ["AccountID", "AccountDescription", "IB"] + month_cols + ["UB"]
    ws.append(headers)
    for _, row in out.iterrows():
        values = [row["AccountID"], row["AccountDescription"], float(row["IB"])]
        for c in month_cols:
            values.append(float(row[c]))
        values.append(float(row["UB"]))
        ws.append(values)

    _auto_cols(ws, [12, 44] + [14] * (len(headers) - 2))
    for r in ws.iter_rows(min_row=2, min_col=3, max_col=2 + len(month_cols) + 1 + 1):
        for cell in r:
            _numfmt(cell)

    wb.save(tb_path)
    return tb_path
