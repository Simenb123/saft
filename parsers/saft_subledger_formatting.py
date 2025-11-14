
# src/app/parsers/saft_subledger_formatting.py
# -*- coding: utf-8 -*-
"""
Formatering av AR/AP subledger-rapporter (Excel). Lav risiko:
- Åpner eksisterende ar_subledger.xlsx / ap_subledger.xlsx og forbedrer kun visning.
- Tallformat: Norsk "Accounting" (# ##0,00) med tusenskiller og komma som desimal.
- Datoformat: dd.mm.yyyy.
- SUM-rad 2 linjer under siste rad for utvalgte ark (Balances, Transactions, *_Balances, *_Transactions).
- Justerer kolonnebredder, fryser topprad og skrur på Autofilter.

Kalles fra saft_gl_monthly.make_gl_monthly(...) for å slippe å endre kall-sti andre steder.
Kan også kjøres direkte:
    python -m app.parsers.saft_subledger_formatting --out <output/csv-folder>
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List
import re
import pandas as pd

ACCOUNTING_FORMAT = '# ##0,00_);(# ##0,00);"-"'
DATE_FORMAT = 'DD.MM.YYYY'  # openpyxl kompatibel

_NUMERIC_HINTS = {
    "IB","PR","UB","Amount","Debit","Credit","Balance","OpenAmount","SumAmount","TaxAmount","GL_TaxAmount",
    "KID","Beløp","Rest","Restbeløp"
}
_DATE_HINTS = {
    "PostingDate","TransactionDate","InvoiceDate","DueDate","Date","BetaltDato","Fakturadato","Forfallsdato"
}
_SUM_SHEETS = {"Balances","Transactions","AR_Balances","AP_Balances","AR_Transactions","AP_Transactions"}

def _is_numeric_header(h: str) -> bool:
    h = str(h).strip()
    if h in _NUMERIC_HINTS: return True
    # fallback: ord/ub-lignende felt
    return bool(re.match(r"^(IB|PR|UB|\d{1,2}|Sum|Beløp|Saldo|MVA|VAT|Base|Tax)", h, re.I))

def _is_date_header(h: str) -> bool:
    h = str(h).strip()
    if h in _DATE_HINTS: return True
    return "date" in h.lower()

def _format_sheet(ws, df: pd.DataFrame, sheet_name: str):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import numbers, Font
    # header indeks
    headers = [cell.value for cell in ws[1]]
    # bestem kolonner
    num_cols_idx: List[int] = []
    date_cols_idx: List[int] = []
    for idx, h in enumerate(headers, start=1):
        if _is_numeric_header(h): num_cols_idx.append(idx)
        if _is_date_header(h): date_cols_idx.append(idx)

    # formater rader
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for ci in num_cols_idx:
            cell = r[ci-1]
            if cell.value is None: 
                continue
            cell.number_format = ACCOUNTING_FORMAT
        for ci in date_cols_idx:
            cell = r[ci-1]
            if cell.value is None: 
                continue
            cell.number_format = DATE_FORMAT

    # kolonnebredder (heuristikk)
    widths = []
    head_len = [len(str(h or "")) for h in headers]
    sample = df.head(500).astype(str).applymap(len).agg("max").fillna(0).astype(int).tolist() if not df.empty else [10]*len(headers)
    for i in range(len(headers)):
        w = max(10, min(60, head_len[i] + 2, 60))
        if i < len(sample):
            w = max(w, min(60, sample[i] + 2))
        widths.append(w)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = float(w)

    # fryse topprad + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # flytt/lag SUM-rad nederst for utvalgte ark
    if sheet_name in _SUM_SHEETS and ws.max_row >= 2:
        # fjern ev. SUM-rad på rad 2 (før data)
        if str(ws.cell(2,1).value).strip().upper() == "SUM":
            ws.delete_rows(2, 1)
        # tom rad + SUM
        row = ws.max_row + 2
        ws.cell(row, 1, "SUM").font = Font(bold=True)
        for ci in num_cols_idx:
            # summer nedover kolonnen (fra rad 2 til siste faktiske rad)
            col_letter = get_column_letter(ci)
            last = ws.max_row
            ws.cell(row, ci, f"=SUM({col_letter}2:{col_letter}{last})").number_format = ACCOUNTING_FORMAT

def _load_df_from_ws(ws) -> pd.DataFrame:
    # trygg henting av header + rader
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    header = [str(h) if h is not None else "" for h in rows[0]]
    body = rows[1:]
    return pd.DataFrame(body, columns=header)

def _format_file(xlsx_path: Path) -> bool:
    import openpyxl
    if not xlsx_path.exists():
        return False
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception:
        return False
    changed = False
    for name in wb.sheetnames:
        ws = wb[name]
        df = _load_df_from_ws(ws)
        _format_sheet(ws, df, name)
        changed = True
    if changed:
        wb.save(xlsx_path)
    return changed

def format_all_subledgers(out_dir: Path | str) -> None:
    out_dir = Path(out_dir)
    excel_dir = out_dir.parent / "excel"
    any_change = False
    for fn in ["ar_subledger.xlsx","ap_subledger.xlsx"]:
        p = excel_dir / fn
        if _format_file(p):
            print(f"[excel] Formaterte {fn} (norsk tall/dato, SUM nederst, filter & frys).")
            any_change = True
    if not any_change:
        print("[excel] Ingen subledger-filer å formatere (hoppet over).")

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Formater AR/AP-subledger Excel-filer i norsk stil.")
    p.add_argument("--out", required=True, help="Output CSV-mappe (samme som for øvrige rapporter)")
    args = p.parse_args()
    format_all_subledgers(args.out)
