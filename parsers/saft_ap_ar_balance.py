# -*- coding: utf-8 -*-
"""
saft_ap_ar_balance.py
---------------------
Lager en samlet arbeidsbok "ap_ar_balance.xlsx" med to ark:
 - AP_Balance: SupplierID, SupplierName, IB_Amount, Movement, UB_Amount
 - AR_Balance: CustomerID, CustomerName, IB_Amount, Movement, UB_Amount

Kilder (i denne rekkefølgen):
  1) csv/ap_subledger.csv   og csv/ar_subledger.csv
  2) ellers ap_subledger.xlsx og ar_subledger.xlsx (leser "flate" tabeller fra første ark)

Kolonnealias:
  ID:    SupplierID | Supplier Id | LeverandørID   /  CustomerID | Customer Id | KundeID
  Navn:  SupplierName | Leverandørnavn             /  CustomerName | Kundenavn
  IB:    IB_Amount | IB Amount | OpeningBalance | Opening | OB Amount | Opening_Amount
  UB:    UB_Amount | UB Amount | ClosingBalance | Closing | CB Amount | Closing_Amount | EndBalance
  MOVE:  Movement | PR Amount | Period Amount | NetChange | Change | PeriodMovement | Amount

Regler:
 - Mangler UB men finnes IB+Movement -> UB = IB + Movement
 - Mangler IB men finnes UB+Movement -> IB = UB - Movement
 - Tall formateres (#,##0.00), header-rad fryses, autofilter settes
 - Summering 2 rader under siste datarad (SUM i IB, Movement, UB)

Dette er en *tilleggsrapport* og endrer ikke eksisterende filer.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple, Optional
import csv

try:
    import xlsxwriter  # for å skrive
except Exception as e:  # pragma: no cover
    raise RuntimeError("xlsxwriter er påkrevd for å skrive ap_ar_balance.xlsx") from e

# Forsøk å lese Excel direkte (fallback når CSV mangler)
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # type: ignore


# --------------------- utils ---------------------

def _sniff_delim(path: Path) -> str:
    text = path.read_text(encoding="utf-8", errors="replace")[:4096]
    if text.count(";") > text.count(","):
        return ";"
    if "\t" in text:
        return "\t"
    if "|" in text:
        return "|"
    return ","


def _read_csv(path: Path) -> List[Dict[str, str]]:
    delim = _sniff_delim(path)
    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f, delimiter=delim)
        for row in r:
            norm: Dict[str, str] = {}
            for k, v in row.items():
                if k is None:
                    continue
                nk = k.strip().lstrip("\ufeff")
                norm[nk] = "" if v is None else str(v).strip()
            rows.append(norm)
    return rows


def _read_first_sheet_xlsx(path: Path) -> List[Dict[str, str]]:
    if load_workbook is None:
        return []
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.worksheets[0]
    headers: List[str] = []
    rows: List[Dict[str, str]] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            headers = [(str(x).strip() if x is not None else "") for x in row]
            first = False
            continue
        rec: Dict[str, str] = {}
        for i, v in enumerate(row):
            key = headers[i] if i < len(headers) else f"col{i+1}"
            rec[key] = "" if v is None else str(v).strip()
        rows.append(rec)
    return rows


def _to_float(s: str) -> float:
    if s is None:
        return 0.0
    t = str(s).strip()
    if not t or t.lower() == "nan":
        return 0.0
    neg = t.startswith("(") and t.endswith(")")
    if neg:
        t = t[1:-1]
    t = t.replace("\xa0", "").replace(" ", "")
    if t.count(",") > 0 and t.count(".") == 0:
        t = t.replace(",", ".")
    try:
        val = float(t)
    except Exception:
        val = 0.0
    return -val if neg else val


def _pick(row: Dict[str, str], keys: Tuple[str, ...]) -> Optional[str]:
    lower = {k.lower(): v for k, v in row.items()}
    for k in keys:
        v = lower.get(k.lower())
        if v not in (None, ""):
            return v
    # substring fallback
    for k in keys:
        kl = k.lower()
        for key, v in lower.items():
            if kl in key and v not in (None, ""):
                return v
    return None


# --------------------- kolonnealias ---------------------

AR_ID_KEYS   = ("CustomerID", "Customer Id", "KundeID", "Kundenr", "KundeNr", "DebitorID")
AR_NAME_KEYS = ("CustomerName", "Kundenavn", "DebitorNavn", "Name")

AP_ID_KEYS   = ("SupplierID", "Supplier Id", "LeverandørID", "LeverandorID", "KreditorID")
AP_NAME_KEYS = ("SupplierName", "Leverandørnavn", "LeverandorNavn", "Name")

IB_KEYS = ("IB_Amount", "IB Amount", "OpeningBalance", "Opening Balance", "Opening", "OB Amount", "Opening_Amount")
UB_KEYS = ("UB_Amount", "UB Amount", "ClosingBalance", "Closing Balance", "Closing", "CB Amount", "Closing_Amount", "EndBalance", "EndingBalance")

# <- her mapper vi PR Amount til Movement (periodens bevegelse)
MOVE_KEYS = ("Movement", "PR Amount", "Period Amount", "NetChange", "Change", "PeriodMovement", "Amount", "Bevegelse", "Netto")


def _normalize_record(row: Dict[str, str], id_keys: Tuple[str, ...], name_keys: Tuple[str, ...]) -> Tuple[str, str, float, float, float]:
    pid = _pick(row, id_keys) or ""
    pname = _pick(row, name_keys) or ""
    ib = _to_float(_pick(row, IB_KEYS))
    mv = _to_float(_pick(row, MOVE_KEYS))
    ub = _to_float(_pick(row, UB_KEYS))

    # Fyll manglende tall hvis mulig
    if ub == 0.0 and (ib != 0.0 or mv != 0.0):
        ub = ib + mv
    elif ib == 0.0 and (ub != 0.0 or mv != 0.0):
        ib = ub - mv

    return pid, pname, ib, mv, ub


def _aggregate(rows: List[Dict[str, str]], id_keys: Tuple[str, ...], name_keys: Tuple[str, ...]) -> List[Tuple[str, str, float, float, float]]:
    agg: Dict[str, Tuple[str, float, float, float]] = {}
    names: Dict[str, str] = {}
    for r in rows:
        pid, pname, ib, mv, ub = _normalize_record(r, id_keys, name_keys)
        if not pid and not pname:
            continue
        old = agg.get(pid or pname, ("", 0.0, 0.0, 0.0))
        _, ib0, mv0, ub0 = old
        agg[pid or pname] = (pid or pname, ib0 + ib, mv0 + mv, ub0 + ub)
        if pid:
            names[pid] = pname or names.get(pid, "")
        else:
            names[pname] = pname
    out: List[Tuple[str, str, float, float, float]] = []
    for k, (pid_or_name, ib, mv, ub) in agg.items():
        name = names.get(k, "")
        if not name and pid_or_name != k:
            name = pid_or_name
        out.append((k, name, ib, mv, ub))
    out.sort(key=lambda x: (x[0] or ""))
    return out


def _load_rows(csv_dir: Path, base: str) -> List[Dict[str, str]]:
    # 1) CSV
    p_csv = csv_dir / f"{base}.csv"
    if p_csv.exists():
        return _read_csv(p_csv)
    # 2) XLSX
    p_xlsx = csv_dir / f"{base}.xlsx"
    if p_xlsx.exists():
        return _read_first_sheet_xlsx(p_xlsx)
    return []


def _write_balance_sheet(wb, sheet_name: str, rows: List[Tuple[str, str, float, float, float]]):
    fmt_head = wb.add_format({"bold": True})
    fmt_amt  = wb.add_format({"num_format": "#,##0.00"})

    ws = wb.add_worksheet(sheet_name)
    ws.write_row(0, 0, [sheet_name.endswith("AP_Balance") and "SupplierID" or "CustomerID",
                        sheet_name.endswith("AP_Balance") and "SupplierName" or "CustomerName",
                        "IB_Amount", "Movement", "UB_Amount"], fmt_head)

    for i, (pid, pname, ib, mv, ub) in enumerate(rows, start=1):
        ws.write(i, 0, pid)
        ws.write(i, 1, pname)
        ws.write_number(i, 2, ib, fmt_amt)
        ws.write_number(i, 3, mv, fmt_amt)
        ws.write_number(i, 4, ub, fmt_amt)

    ws.freeze_panes(1, 0)
    ws.autofilter(0, 0, max(1, len(rows)), 4)
    ws.set_column(0, 0, 14)
    ws.set_column(1, 1, 42)
    ws.set_column(2, 4, 18)

    # Summelinje: 2 rader under siste data-rad
    last = len(rows)  # 1-basert data slutter på rad = last
    sum_row = last + 2  # tom rad + summerad
    if last >= 1:
        ws.write(sum_row, 1, "SUM")
        # Formler (Excel-radindekser er 1-basert i formler)
        start_r = 2
        end_r = last + 1
        ws.write_formula(sum_row, 2, f"=SUM(C{start_r}:C{end_r})", fmt_amt)
        ws.write_formula(sum_row, 3, f"=SUM(D{start_r}:D{end_r})", fmt_amt)
        ws.write_formula(sum_row, 4, f"=SUM(E{start_r}:E{end_r})", fmt_amt)


def make_ap_ar_balance(csv_dir: Path) -> Path:
    """
    Leser subledger-rader, aggregerer pr. partner og skriver ap_ar_balance.xlsx med:
      - AP_Balance
      - AR_Balance
    """
    csv_dir = Path(csv_dir)

    ap_rows = _load_rows(csv_dir, "ap_subledger")
    ar_rows = _load_rows(csv_dir, "ar_subledger")

    ap_agg = _aggregate(ap_rows, AP_ID_KEYS, AP_NAME_KEYS) if ap_rows else []
    ar_agg = _aggregate(ar_rows, AR_ID_KEYS, AR_NAME_KEYS) if ar_rows else []

    xlsx_path = csv_dir / "ap_ar_balance.xlsx"
    wb = xlsxwriter.Workbook(str(xlsx_path))

    # Viktig: skriv AR først hvis du vil ha AR_Balance som første ark; vi lar AP først siden filnavn antyder AP/AR.
    if ap_agg:
        _write_balance_sheet(wb, "AP_Balance", ap_agg)
    if ar_agg:
        _write_balance_sheet(wb, "AR_Balance", ar_agg)

    wb.close()
    return xlsx_path
